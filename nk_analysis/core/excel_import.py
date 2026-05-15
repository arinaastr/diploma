import os
import openpyxl
import pandas as pd

def load_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    result = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) >= 2:
            result[name] = pd.DataFrame(rows[1:], columns=rows[0])
        elif len(rows) == 1:
            result[name] = pd.DataFrame(columns=rows[0])
    return result

def parse_stvol(df):
    records = []
    cols = [str(c) for c in df.columns]

    horiz_col = df.columns[0]

    sides_uk = [c for c in cols if c.endswith('_УК') or c.endswith('_ук')]
    sides_mo = [c for c in cols if c.endswith('_МО') or c.endswith('_мо')]

    if sides_uk and sides_mo:
        for uk_col in sides_uk:
            side = uk_col.split('_')[0]
            mo_col = side + '_МО'
            if mo_col not in cols:
                mo_col = side + '_мо'
            if mo_col not in cols:
                mo_col = None
                for mc in sides_mo:
                    if mc.split('_')[0] == side:
                        mo_col = mc
                        break

            for _, row in df.iterrows():
                horiz = row[horiz_col]
                v_val = row[uk_col]
                f_val = row[mo_col] if mo_col and mo_col in df.columns else None
                if v_val is not None:
                    records.append({
                        "Горизонт": horiz,
                        "Сторона": side,
                        "V": v_val,
                        "f_МО": f_val,
                    })
    else:
        v_cols  = [c for c in cols if 'v' in c.lower() or 'скор' in c.lower() or 'ук' in c.lower()]
        f_cols  = [c for c in cols if 'f' in c.lower() or 'прочн' in c.lower() or 'мо' in c.lower()]
        sides_list = ['С', 'Ю', 'В', 'З']
        if v_cols and f_cols:
            for _, row in df.iterrows():
                horiz = row[horiz_col]
                for vc, fc in zip(v_cols, f_cols):
                    side = vc.split('_')[0] if '_' in vc else '—'
                    records.append({
                        "Горизонт": horiz, "Сторона": side,
                        "V": row[vc], "f_МО": row.get(fc),
                    })

    result = pd.DataFrame(records)
    if len(result):
        result["V"]    = pd.to_numeric(result["V"],    errors="coerce")
        result["f_МО"] = pd.to_numeric(result["f_МО"], errors="coerce")
    return result

def parse_ne_stvol(df):
    rename = {}
    for col in df.columns:
        cs = str(col).strip().lower()
        if cs in ('v', 'скорость', 'скорость ультразвука', 'скорость узк', 'v, м/с'):
            rename[col] = 'V'
        elif any(k in cs for k in ('прочность мо', 'прочность по мо', 'f_мо', 'прочн', 'f, мпа', 'мо')):
            rename[col] = 'f_МО'
        elif any(k in cs for k in ('участок', 'номер участка', 'конструкц', 'элем', 'наимен')):
            rename[col] = 'Участок'

    result = df.rename(columns=rename).copy()
    for col in ['V', 'f_МО']:
        if col in result:
            result[col] = pd.to_numeric(result[col], errors='coerce')
    if 'Участок' not in result.columns:
        if len(df.columns) >= 2:
            result['Участок'] = df.iloc[:, 1].astype(str)
        else:
            result['Участок'] = [f'Конструкция {i+1}' for i in range(len(result))]
    return result

def extract_pairs(df):
    if df is None or len(df) == 0:
        return pd.DataFrame(columns=['V', 'f'])
    if 'V' not in df.columns or 'f_МО' not in df.columns:
        return pd.DataFrame(columns=['V', 'f'])
    pairs = df.dropna(subset=['V', 'f_МО']).copy()
    pairs = pairs.rename(columns={'f_МО': 'f'})[['V', 'f']]
    pairs = pairs[(pairs['V'] > 100) & (pairs['f'] > 0) & (pairs['f'] < 200)]
    return pairs.reset_index(drop=True)

def load_all_files(paths):
    all_s, all_n, errors = [], [], []
    for p in paths:
        fn = os.path.basename(p)
        try:
            sheets = load_xlsx(p)
            sheet_upper = {k.upper().replace(' ', '_'): k for k in sheets}
            if 'СТВОЛ' in sheet_upper:
                df = parse_stvol(sheets[sheet_upper['СТВОЛ']])
                df['Файл'] = fn
                all_s.append(df)
            if 'НЕ_СТВОЛ' in sheet_upper or 'НЕ СТВОЛ' in sheet_upper:
                key = sheet_upper.get('НЕ_СТВОЛ') or sheet_upper.get('НЕ СТВОЛ')
                df = parse_ne_stvol(sheets[key])
                df['Файл'] = fn
                all_n.append(df)
        except Exception as e:
            errors.append(f'{fn}: {e}')
    ds = pd.concat(all_s, ignore_index=True) if all_s else pd.DataFrame()
    dn = pd.concat(all_n, ignore_index=True) if all_n else pd.DataFrame()
    return ds, dn, extract_pairs(ds), extract_pairs(dn), errors
